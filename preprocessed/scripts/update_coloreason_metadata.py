#!/usr/bin/env python3
"""Apply the canonical ColoReason field schema to a metadata workbook."""

import argparse

from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(__file__).resolve().parents[1] / "assets" / "ColoReason.xlsx"

FIELD_RENAMES = {
    "肿瘤长径(cm": "肿瘤长径(cm)",
}

HISTORY_FIELDS = {
    "有无意识障碍",
    "有无高血压",
    "有无糖尿病",
    "有无冠心病",
    "有无高尿酸血症",
    "有无高血脂症",
    "有无病毒性肝炎",
    "有无血吸虫病",
}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()
    workbook = load_workbook(args.workbook)
    sheet = workbook["In-house Data"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}

    changed = []
    for row in range(2, sheet.max_row + 1):
        field_cell = sheet.cell(row, headers["结构化字段"])
        field = field_cell.value
        if field in FIELD_RENAMES:
            new_value = FIELD_RENAMES[field]
            field_cell.value = new_value
            changed.append((row, "结构化字段", field, new_value))
            field = new_value

        if field in HISTORY_FIELDS:
            for column, new_value in (
                ("出处数据表（拼音）", "final_yisuwushi"),
                ("出处数据表（中文）", "一诉五史"),
            ):
                cell = sheet.cell(row, headers[column])
                old_value = cell.value
                if old_value != new_value:
                    cell.value = new_value
                    changed.append((row, column, old_value, new_value))

    workbook.save(args.workbook)
    print(f"updated={args.workbook} changes={len(changed)}")
    for item in changed:
        print(item)


if __name__ == "__main__":
    main()
