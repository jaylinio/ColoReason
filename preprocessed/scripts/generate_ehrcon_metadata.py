#!/usr/bin/env python3
"""Build EHRCon field metadata from preprocessed statistics and a workbook template."""

from __future__ import annotations

import argparse
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_NAMES = {
    "discharge": ("Discharge Summary", "出院小结"),
    "physician": ("Physician Note", "医师记录"),
    "nursing": ("Nursing Note", "护理记录"),
}

TABLE_NAMES = {
    "chartevents": "监测事件",
    "d_icd_diagnoses": "ICD诊断字典",
    "d_icd_procedures": "ICD操作字典",
    "d_items": "监测项目字典",
    "d_labitems": "检验项目字典",
    "inputevents_cv": "CareVue输入事件",
    "inputevents_mv": "MetaVision输入事件",
    "labevents": "检验事件",
    "microbiologyevents": "微生物事件",
    "outputevents": "输出事件",
    "prescriptions": "处方",
}

FIELD_NAMES = {
    "amount": "输入量",
    "amountuom": "输入量单位",
    "chartdate": "记录日期",
    "charttime": "记录时间",
    "dose_unit_rx": "处方剂量单位",
    "dose_val_rx": "处方剂量值",
    "drug": "药物名称",
    "fluid": "标本/体液类型",
    "form_unit_disp": "发放剂型单位",
    "label": "项目名称",
    "long_title": "完整名称",
    "org_name": "微生物名称",
    "originalroute": "原始给药途径",
    "rate": "输入速率",
    "rateuom": "输入速率单位",
    "route": "给药途径",
    "short_title": "简称",
    "spec_type_desc": "标本类型描述",
    "value": "记录值",
    "valuenum": "数值结果",
    "valueuom": "结果单位",
}

ENGLISH_FIELD_NAMES = {
    "amount": "Amount",
    "amountuom": "Amount unit",
    "chartdate": "Chart date",
    "charttime": "Chart time",
    "dose_unit_rx": "Prescription dose unit",
    "dose_val_rx": "Prescription dose value",
    "drug": "Drug name",
    "fluid": "Specimen/fluid type",
    "form_unit_disp": "Dispensed form unit",
    "label": "Item label",
    "long_title": "Long title",
    "org_name": "Organism name",
    "originalroute": "Original administration route",
    "rate": "Infusion rate",
    "rateuom": "Infusion rate unit",
    "route": "Administration route",
    "short_title": "Short title",
    "spec_type_desc": "Specimen type description",
    "value": "Recorded value",
    "valuenum": "Numeric value",
    "valueuom": "Value unit",
}

HEADERS = [
    "结构化字段",
    "英文翻译",
    "层次",
    "评估方法",
    "是否需要推理",
    "推理线索",
    "出处数据表（拼音）",
    "出处数据表（中文）",
    "出处字段（拼音）",
    "出处字段（中文）",
    "非结构化字段",
    "出处数据表",
    "Gold字段出现次数",
    "Type 1次数",
    "Type 2次数",
    "唯一Entity数",
    "Gold Note数",
    "字段值示例",
    "筛选规则",
    "数据划分",
]


def field_protocol(field_name: str) -> tuple[str, str, str]:
    field = field_name.rsplit(".", 1)[1]
    if field in {"valuenum", "value", "amount", "rate", "dose_val_rx"}:
        return (
            "Fuzzy",
            "是",
            "从临床文本定位对应实体及数值，并结合时间窗、单位和项目名称，与MIMIC-III结构化记录核对。",
        )
    if field in {"charttime", "chartdate"}:
        return (
            "Fuzzy",
            "是",
            "解析文本中的标准时间、叙述性时间或缺失时间；出院小结使用住院期，医师/护理记录使用记录日前后1天作为默认时间窗。",
        )
    if field.endswith("uom") or field in {"dose_unit_rx", "form_unit_disp"}:
        return (
            "Semantics",
            "是",
            "抽取实体的单位表达并进行缩写、大小写和等价单位归一，再与结构化字段核对。",
        )
    return (
        "Semantics",
        "是",
        "识别文本实体及其同义词/缩写，定位对应MIMIC-III项目或事件记录，并判断结构化记录是否支持该实体。",
    )


def copy_style(src, dst) -> None:
    dst._style = copy(src._style)
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def build_workbook(template_path: Path, stats_path: Path, output_path: Path) -> None:
    with stats_path.open(encoding="utf-8") as fh:
        payload = json.load(fh)

    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = "EHRCon Data"

    header_style = [copy(cell._style) for cell in ws[1][:12]]
    group_style = [copy(cell._style) for cell in ws[3][:12]]
    data_style = [copy(cell._style) for cell in ws[4][:12]]

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    if ws.max_column > len(HEADERS):
        ws.delete_cols(len(HEADERS) + 1, ws.max_column - len(HEADERS))

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, col_idx, header)
        cell._style = copy(header_style[min(col_idx, 12) - 1])
        cell.font = copy(cell.font)
        cell.font = Font(name="Arial", size=cell.font.sz or 11, bold=True, color=cell.font.color)
        cell.fill = PatternFill("solid", fgColor="E5F6FF")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    manifest = payload["manifest"]
    ws.cell(2, 1, "外部测试任务：从临床Note提取MIMIC-III结构化键值；仅使用entity-level完全一致标注作为Gold")
    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(2, col_idx)._style = copy(group_style[min(col_idx, 12) - 1])
        ws.cell(2, col_idx).font = Font(name="Arial", bold=True)

    row_idx = 3
    for source in ("discharge", "physician", "nursing"):
        source_en, source_zh = SOURCE_NAMES[source]
        rows = [item for item in payload["fields"] if item["note_type"] == source]
        source_stat = manifest["source_counts"][source]
        ws.cell(row_idx, 1, f"{source_en}（{source_zh}）")
        ws.cell(
            row_idx,
            2,
            f"{source_stat['gold_entities']} gold entities; {source_stat['gold_notes']} notes; {len(rows)} structured fields",
        )
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row_idx, col_idx)._style = copy(group_style[min(col_idx, 12) - 1])
            ws.cell(row_idx, col_idx).font = Font(name="Arial", bold=True)
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="D9EAD3")
        row_idx += 1

        for item in rows:
            full_field = item["field"]
            table, field = full_field.split(".", 1)
            evaluation, reasoning, clue = field_protocol(full_field)
            values = [
                f"{TABLE_NAMES.get(table, table)}.{FIELD_NAMES.get(field, field)}",
                f"{table}.{ENGLISH_FIELD_NAMES.get(field, field)}",
                "Entity-level",
                evaluation,
                reasoning,
                clue,
                table,
                TABLE_NAMES.get(table, table),
                field,
                FIELD_NAMES.get(field, field),
                "text",
                f"{source_en}（{source_zh}）",
                item["gold_occurrences"],
                item["type1"],
                item["type2"],
                item["unique_entities"],
                item["gold_notes"],
                "；".join(item["values"]),
                "Type 1/2 + consistency + errors为空/NaN + text/fields非空；删除占位值",
                "test + valid",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row_idx, col_idx, value)
                cell._style = copy(data_style[min(col_idx, 12) - 1])
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx == 18:
                    cell.number_format = "@"
            row_idx += 1

    if ws.max_row >= row_idx:
        ws.delete_rows(row_idx, ws.max_row - row_idx + 1)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{row_idx - 1}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 26

    widths = [34, 36, 14, 13, 14, 72, 22, 24, 24, 24, 18, 30, 18, 13, 13, 15, 14, 46, 52, 15]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--stats", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    build_workbook(args.template, args.stats, args.output)


if __name__ == "__main__":
    main()
